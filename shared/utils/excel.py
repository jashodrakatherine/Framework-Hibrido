from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook


def read_excel(path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


def write_excel(path: str, data: List[Dict[str, Any]], sheet_name: str = "Sheet1") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if data:
        headers = list(data[0].keys())
        ws.append(headers)
        for row in data:
            ws.append([row.get(h) for h in headers])

    wb.save(path)
